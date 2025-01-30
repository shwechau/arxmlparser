"""
    Author : Shweta Chauhan

    This tool is to extract arxml files in an excel sheet, sort the messages based on the ID's and then calculate message delay per CAN cluster

"""
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl import load_workbook
import re
import csv
import sys
import json
import os
import argparse
import pandas


class ArxmlParser:
    def __init__(self, filepath="EthCluster11_L.arxml"):
        self.file_path = filepath
        self.tree = ET.parse(filepath)
        self.root = self.tree.getroot()

    def parseXML(self, search):
        items = self.tree.findall('.//{http://autosar.org/schema/r4.0}AR-PACKAGE')
        for arpackage in self.tree.findall('.//{http://autosar.org/schema/r4.0}AR-PACKAGE'):
            arpkcg = arpackage.findall('./{http://autosar.org/schema/r4.0}SHORT-NAME')[0].text
            if arpkcg == search:
                topPackage = arpackage
                return topPackage

    def getclusterdata(self,topPackage):
        all_channel = []
        for item in topPackage.findall('.//{http://autosar.org/schema/r4.0}ELEMENTS'):
            for CANChannel in item.findall('.//{http://autosar.org/schema/r4.0}CAN-PHYSICAL-CHANNEL'):
                all_channel.append(CANChannel)
        return all_channel

    def getpdudata(self, pdu, element):
        self.pdu_dict = {}
        for signalPDU in element.findall('.//{http://autosar.org/schema/r4.0}'+pdu):
            PDU_name= signalPDU.findall('./{http://autosar.org/schema/r4.0}SHORT-NAME')[0].text
            PDU_length_list = signalPDU.findall('./{http://autosar.org/schema/r4.0}LENGTH')
            pdu_cycle_time = 0
            for i_pdu_timing_spec in signalPDU.findall('./{http://autosar.org/schema/r4.0}I-PDU-TIMING-SPECIFICATIONS'):
                for i_pdu_timing in i_pdu_timing_spec.findall('./{http://autosar.org/schema/r4.0}I-PDU-TIMING'):
                    for trans_mode_decl in i_pdu_timing.findall('./{http://autosar.org/schema/r4.0}TRANSMISSION-MODE-DECLARATION'):
                        for trans_mode_timing in trans_mode_decl.findall('./{http://autosar.org/schema/r4.0}TRANSMISSION-MODE-TRUE-TIMING'):
                            for cyclic_time in trans_mode_timing.findall('./{http://autosar.org/schema/r4.0}CYCLIC-TIMING'):
                                for time_period in cyclic_time.findall('./{http://autosar.org/schema/r4.0}TIME-PERIOD'):
                                    pdu_cycle_time = float(time_period.findall('./{http://autosar.org/schema/r4.0}VALUE')[0].text)
            if len(PDU_length_list) == 0:
                continue
            else:
                self.pdu_dict[PDU_name] = [PDU_name, int(PDU_length_list[0].text), pdu_cycle_time]


    def getframedata(self, channel_list):
        channel_dict = {}
        for channel in channel_list:
            channel_name = channel.findall('./{http://autosar.org/schema/r4.0}SHORT-NAME')[0].text
            self.frame_list = []
            self.frame_id = []
            self.frames = []
            self.pdu_length_list = []
            self.cycle_time_list = []
            PDUpckg = self.parseXML('PDUs')
            for element in PDUpckg.findall('.//{http://autosar.org/schema/r4.0}ELEMENTS'):
                # self.getpdudata('DCM-I-PDU', element)
                self.getpdudata('GENERAL-PURPOSE-PDU', element)
                self.getpdudata('I-SIGNAL-I-PDU', element)
            for frame_trigger in channel.findall('.//{http://autosar.org/schema/r4.0}CAN-FRAME-TRIGGERING'):
                frame_name = frame_trigger.findall('./{http://autosar.org/schema/r4.0}SHORT-NAME')[0].text
                try:
                    frameid = frame_trigger.findall('./{http://autosar.org/schema/r4.0}IDENTIFIER')[0].text
                    if frame_name in self.pdu_dict:
                        self.pdu_length_list.append(self.pdu_dict[frame_name][1])
                        self.cycle_time_list.append(self.pdu_dict[frame_name][2])
                        self.frame_list.append(frame_name)
                        self.frame_id.append(int(frameid))
                except:
                    print('error in the id -----> ', frame_name)
                    self.frame_list.append(frame_name)
                    self.frame_id.append(-1)
                    self.pdu_length_list.append(-1)
                    self.cycle_time_list.append(-1)
                channel_dict[channel_name] = self.frame_list, self.frame_id, self.pdu_length_list, self.cycle_time_list
        return channel_dict

    def csvinfo(self, channel_dict):

        wb=Workbook()
        filepath="C:\\Users\\ShwetaChauhan\\IdeaProjects\\ARXML\\Eth21abc.xlsx"
        wb.save(filepath)
        frame_list = self.frame_list
        frame_id = self.frame_id
        for item in channel_dict:
            ws = wb.create_sheet(title=item)
            ws = wb[item]
            i=1
            ws.cell(row=1, column=1).value = 'FRAME'
            ws.cell(row=1, column=2).value = 'ID'
            ws.cell(row=1, column=3).value = 'Frame_Length'
            ws.cell(row=1, column=4).value = 'Cycle_Time'
            frame_identifiers_length = channel_dict[item]
            print(frame_identifiers_length)
            frame_list_zipped = list(zip(frame_identifiers_length[0],frame_identifiers_length[1], frame_identifiers_length[2], frame_identifiers_length[3]))
            print('checkzipped')
            print(frame_list_zipped)
            for row in range(0,len(frame_list_zipped)):
                tuple_list = list(frame_list_zipped[row])
                for column in range(0,len(tuple_list)):
                    ws.cell(row = row+2, column = column+1).value = tuple_list[column]
            ws.auto_filter.ref = ws.dimensions
        wb.save(filepath)



if __name__ == '__main__':
    parser = ArxmlParser()
    topology = parser.parseXML('Topology')
    channel_list = parser.getclusterdata(topology)
    channel_dict = parser.getframedata(channel_list)
    print(channel_dict)
    parser.csvinfo(channel_dict)






